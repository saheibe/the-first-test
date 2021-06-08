
import os  # 引入工具函数库 os 
import re  # 正则表达式库
from openpyxl import Workbook    #excel库，需要安装 pip install openpyxl
path = os.path.abspath(os.path.dirname(__file__)) #获取当前脚本所在目录


# 时间，原子a，原子b，距离 

TYPE='3' #仅类型3原子之间的连接
DICT={} #所有原子id对应原子类型的字典
TIME=0   #时间，单位s
STOPTIME= 55555  #只分析n秒内的数据
RESULTLIST=[]   #用来保存结果集


# 函数：用来分析一行数据，将分析出来的结果，存入RESULTLIST 。 参数arr: 某行数据转成的list ，distinctList：该秒已经记录的原子连接组合（用来去重） ，time：是第n秒的记录
def getSameLink( arr, distinctList ,time ):
    if arr[1]!=TYPE:
        return #类型不对
    countNum=  int(arr[2])
    if countNum <1:
        return #无连接的独立原子
    for i in range(1 ,countNum+1 ):
        yuanzib=arr[2+i]
        yuanzia=arr[0]
        if  DICT[yuanzib]!=TYPE :
            continue #连接的这个原子不是对应类型
        key = '' #将连接的两个原子id从小到达排列，组合成唯一key ，存入list，记录写入状态。
        if (int(yuanzia) < int(yuanzib)):
            key =yuanzia + '-'+yuanzib
        else : 
            key = yuanzib+  '-'+yuanzia 
        if distinctList.count( key )<1 : # 已经记录过该原子连接组合，就不再记录
           distinctList.append(key)
           templl= [ int( TIME) ,int( yuanzia) ,int(yuanzib) , key , float(arr[2+i+countNum+1]) ]
           print(templl)
           RESULTLIST.append(templl ) 


print("脚本所在目录："+ path)  #打印脚本所在目录
inputFile = open ( path +"/bonds.reaxc" ,"r")  # 以读取模式打开文件 ，用于读取 数据源文件 

line = inputFile.readline()  #从上次读取结束的位置读取下一行，第一次读所以在是第一行
while line:  #循环体，读不到行就结束循环。 python语法比较特殊的点， 用缩进量体现语句块结束位置 。相同缩进量构成一个开始和结束 
    matchObj = re.match( '#\sTimestep\s(\d+)' ,line )
    if matchObj :
        time = matchObj.group(1)
        #print('time :' + time )
        if time !='0':
            break
        line = inputFile.readline() 
        continue
    elif  re.match('#.*' ,line ):
        line = inputFile.readline() 
        continue
    #print(line)
    temp = line.split() #用空格分隔本行文本得到一个数组。   
    if len(temp) > 1 : #防止数组索引越界
        DICT[temp[0]] =temp[1]
        #print ("打印本行内容:第一列"+ temp[0] + "第二列 "+temp[1]) 

        
    line = inputFile.readline()  #从上次读取结束的位置读取下一行   。 本行是循环代码的结束位置  ，因为下一行没有缩进了 
print(DICT)

inputFile.seek(0) #返回文件头
line = inputFile.readline()  #从上次读取结束的位置读取下一行，第一次读所以在是第一行
distinctList=[]
while line:  #循环体，读不到行就结束循环。 python语法比较特殊的点， 用缩进量体现语句块结束位置 。相同缩进量构成一个开始和结束 
    matchObj = re.match( '#\sTimestep\s(\d+)' ,line )
    
    if matchObj :
        time = matchObj.group(1)        
        TIME =int(int(time)/1000)
        print('time :' + str( TIME) +'===================================================================' )
        if TIME==STOPTIME : 
            break 
        line = inputFile.readline() 
        distinctList.clear()
        continue
    elif  re.match('#.*' ,line ):
        line = inputFile.readline() 
        continue
    #print(line)
    temp = line.split() #用空格分隔本行文本得到一个数组。   
    if len(temp) > 1 : #防止数组索引越界
        getSameLink( temp ,distinctList ,TIME )
      
        
    line = inputFile.readline()  #从上次读取结束的位置读取下一行   。 本行是循环代码的结束位置  ，因为下一行没有缩进了 


inputFile.close() #关闭文件

print( 'Begin create Excel,wait')
wb = Workbook()
ws1 = wb.create_sheet("yuanzi")
for k, v in DICT.items()  : 
    ws1.append( [int(k),int(v)] )  

for x in RESULTLIST :
    index =int(x[0]/200) #每200秒创建一个sheelt
    sheetname=str(index*200)+'-'+str(index*200+200)
    if sheetname not in wb.sheetnames:    
        ws3 = wb.create_sheet(sheetname ,  index )
    ws3.append(x)
    

print( 'Please wait a moment')
wb.save(path+'/balances'+ TYPE+'.xlsx')

print( 'Successful completion ')


